"""Generic Excel import/export for ALL CMS entities.

Endpoints (all admin-only):
  GET  /api/excel/{entity}/template
  POST /api/excel/{entity}/export   (body: {ids?: [str]})
  POST /api/excel/{entity}/import   (multipart file)

Entities: products, providers, customers, categories, users
(Orders already implemented in routers_excel.py; this complements it.)
"""
import io
import uuid
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import db
from auth import require_permission, hash_password, validate_password, ALL_PERMISSIONS

router = APIRouter(prefix="/api/excel", tags=["excel"])


# ---------- styling helpers ----------
def _style_header(ws, columns):
    header_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    header_font = Font(color="C5A059", bold=True, size=10)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 26


def _build_xlsx(sheet_name: str, columns: list, rows: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _style_header(ws, columns)
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    for col_idx in range(1, len(columns) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(14, min(40, len(columns[col_idx - 1]) + 4))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(data: bytes, filename: str) -> Response:
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _read_xlsx(file_data: bytes):
    try:
        wb = load_workbook(io.BytesIO(file_data))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido")
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    headers = [str(h or "").strip() for h in rows[0]]
    data_rows = [dict(zip(headers, r)) for r in rows[1:] if any(v is not None and v != "" for v in r)]
    return headers, data_rows


# ============ PRODUCTS ============
PRODUCT_COLS = [
    "SKU", "Nombre", "Descripción corta", "Descripción larga",
    "Precio (IVA incl.)", "Precio comparación", "IVA %", "Stock",
    "Umbral stock bajo", "Peso (g)", "Categoría (nombre)", "Proveedor (nombre)",
    "Origen", "Curación (meses)", "Raza", "Alimentación",
    "Tags (coma)", "Activo (SI/NO)", "Destacado (SI/NO)", "Alta",
]


def _bool(v) -> bool:
    if isinstance(v, bool): return v
    s = str(v or "").strip().lower()
    return s in ("si", "sí", "yes", "y", "true", "1", "x")


def _product_to_row(p, categories_by_id, providers_by_id):
    return [
        p.get("sku", ""), p.get("name", ""), p.get("description", ""),
        p.get("long_description", ""),
        float(p.get("price", 0)),
        float(p.get("compare_at_price") or 0) if p.get("compare_at_price") else "",
        int(p.get("vat_rate", 10)),
        int(p.get("stock", 0)),
        int(p.get("low_stock_threshold", 5)),
        int(p.get("weight_grams") or 0) if p.get("weight_grams") else "",
        categories_by_id.get(p.get("category_id"), "") if p.get("category_id") else "",
        providers_by_id.get(p.get("provider_id"), "") if p.get("provider_id") else "",
        p.get("origin", ""),
        int(p.get("curing_months") or 0) if p.get("curing_months") else "",
        p.get("breed", ""), p.get("feed", ""),
        ", ".join(p.get("tags", []) or []),
        "SI" if p.get("is_active", True) else "NO",
        "SI" if p.get("is_featured") else "NO",
        (p.get("created_at") or "")[:10],
    ]


@router.get("/products/template")
async def products_template(_=Depends(require_permission("products.read"))):
    sample = [{
        "sku": "JIB-EJE-001", "name": "Jamón ejemplo", "description": "Descripción breve",
        "long_description": "Descripción larga del producto",
        "price": 49.90, "compare_at_price": 0, "vat_rate": 10, "stock": 10,
        "low_stock_threshold": 3, "weight_grams": 500,
        "category_id": None, "provider_id": None,
        "origin": "Sierra Norte de Sevilla", "curing_months": 24,
        "breed": "Ibérico", "feed": "Bellota",
        "tags": ["jamones", "bellota"], "is_active": True, "is_featured": False,
        "created_at": "",
    }]
    data = _build_xlsx("Productos", PRODUCT_COLS, [_product_to_row(s, {}, {}) for s in sample])
    return _xlsx_response(data, "plantilla_productos.xlsx")


class IdsIn(BaseModel):
    ids: Optional[List[str]] = None


@router.post("/products/export")
async def products_export(payload: IdsIn, _=Depends(require_permission("products.read"))):
    cats = await db.categories.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    provs = await db.providers.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    cats_by_id = {c["id"]: c["name"] for c in cats}
    provs_by_id = {p["id"]: p["name"] for p in provs}

    query = {"id": {"$in": payload.ids}} if payload.ids else {}
    products = await db.products.find(query, {"_id": 0}).to_list(2000)
    if not products:
        raise HTTPException(status_code=404, detail="Sin productos para exportar")
    rows = [_product_to_row(p, cats_by_id, provs_by_id) for p in products]
    return _xlsx_response(_build_xlsx("Productos", PRODUCT_COLS, rows), "productos_export.xlsx")


@router.post("/products/import")
async def products_import(file: UploadFile = File(...), _=Depends(require_permission("products.write"))):
    data = await file.read()
    _, rows = _read_xlsx(data)
    cats = {c["name"].lower(): c["id"] async for c in db.categories.find({}, {"_id": 0, "id": 1, "name": 1})}
    provs = {p["name"].lower(): p["id"] async for p in db.providers.find({}, {"_id": 0, "id": 1, "name": 1})}
    imported = updated = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    for idx, r in enumerate(rows, start=2):
        try:
            sku = str(r.get("SKU") or "").strip()
            if not sku: continue
            cat_id = cats.get(str(r.get("Categoría (nombre)") or "").strip().lower())
            prov_id = provs.get(str(r.get("Proveedor (nombre)") or "").strip().lower())
            tags_str = str(r.get("Tags (coma)") or "")
            tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            doc = {
                "name": str(r.get("Nombre") or sku),
                "description": str(r.get("Descripción corta") or ""),
                "long_description": str(r.get("Descripción larga") or ""),
                "price": float(r.get("Precio (IVA incl.)") or 0),
                "compare_at_price": float(r.get("Precio comparación")) if r.get("Precio comparación") else None,
                "vat_rate": int(r.get("IVA %") or 10),
                "stock": int(r.get("Stock") or 0),
                "low_stock_threshold": int(r.get("Umbral stock bajo") or 5),
                "weight_grams": int(r.get("Peso (g)")) if r.get("Peso (g)") else None,
                "category_id": cat_id,
                "provider_id": prov_id,
                "origin": str(r.get("Origen") or ""),
                "curing_months": int(r.get("Curación (meses)")) if r.get("Curación (meses)") else None,
                "breed": str(r.get("Raza") or ""),
                "feed": str(r.get("Alimentación") or ""),
                "tags": tags,
                "is_active": _bool(r.get("Activo (SI/NO)")) if r.get("Activo (SI/NO)") is not None else True,
                "is_featured": _bool(r.get("Destacado (SI/NO)")),
                "updated_at": now,
            }
            existing = await db.products.find_one({"sku": sku})
            if existing:
                await db.products.update_one({"sku": sku}, {"$set": doc})
                updated += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["sku"] = sku
                doc["images"] = []
                doc["created_at"] = now
                await db.products.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {e}")
    return {"imported": imported, "updated": updated, "errors": errors}


# ============ PROVIDERS ============
PROVIDER_COLS = ["Nombre", "Empresa", "Contacto", "Email", "Teléfono",
                 "Dirección", "Ciudad", "CP", "País", "CIF/NIF",
                 "Web", "Condiciones pago", "Tags (coma)", "Activo (SI/NO)", "Notas"]


def _provider_row(p):
    return [
        p.get("name", ""), p.get("company", ""), p.get("contact_name", ""),
        p.get("email", ""), p.get("phone", ""), p.get("address", ""),
        p.get("city", ""), p.get("postal_code", ""), p.get("country", "España"),
        p.get("tax_id", ""), p.get("website", ""), p.get("payment_terms", ""),
        ", ".join(p.get("tags", []) or []),
        "SI" if p.get("is_active", True) else "NO",
        p.get("notes", ""),
    ]


@router.get("/providers/template")
async def providers_template(_=Depends(require_permission("products.read"))):
    sample = [_provider_row({
        "name": "Proveedor Ejemplo", "email": "proveedor@example.com",
        "phone": "+34 600 000 000", "country": "España", "is_active": True,
    })]
    return _xlsx_response(_build_xlsx("Proveedores", PROVIDER_COLS, sample), "plantilla_proveedores.xlsx")


@router.post("/providers/export")
async def providers_export(payload: IdsIn, _=Depends(require_permission("products.read"))):
    query = {"id": {"$in": payload.ids}} if payload.ids else {}
    items = await db.providers.find(query, {"_id": 0}).to_list(2000)
    if not items:
        raise HTTPException(status_code=404, detail="Sin proveedores")
    return _xlsx_response(_build_xlsx("Proveedores", PROVIDER_COLS, [_provider_row(p) for p in items]), "proveedores_export.xlsx")


@router.post("/providers/import")
async def providers_import(file: UploadFile = File(...), _=Depends(require_permission("products.write"))):
    data = await file.read()
    _, rows = _read_xlsx(data)
    imported = updated = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    for idx, r in enumerate(rows, start=2):
        try:
            email = str(r.get("Email") or "").strip().lower()
            if not email: continue
            tags = [t.strip() for t in str(r.get("Tags (coma)") or "").split(",") if t.strip()]
            doc = {
                "name": str(r.get("Nombre") or email),
                "company": str(r.get("Empresa") or ""),
                "contact_name": str(r.get("Contacto") or ""),
                "phone": str(r.get("Teléfono") or ""),
                "address": str(r.get("Dirección") or ""),
                "city": str(r.get("Ciudad") or ""),
                "postal_code": str(r.get("CP") or ""),
                "country": str(r.get("País") or "España"),
                "tax_id": str(r.get("CIF/NIF") or ""),
                "website": str(r.get("Web") or ""),
                "payment_terms": str(r.get("Condiciones pago") or ""),
                "tags": tags,
                "is_active": _bool(r.get("Activo (SI/NO)")) if r.get("Activo (SI/NO)") is not None else True,
                "notes": str(r.get("Notas") or ""),
                "updated_at": now,
            }
            existing = await db.providers.find_one({"email": email})
            if existing:
                await db.providers.update_one({"email": email}, {"$set": doc})
                updated += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["email"] = email
                doc["created_at"] = now
                await db.providers.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {e}")
    return {"imported": imported, "updated": updated, "errors": errors}


# ============ CATEGORIES ============
CATEGORY_COLS = ["Slug", "Nombre", "Descripción"]


@router.get("/categories/template")
async def categories_template(_=Depends(require_permission("products.read"))):
    sample = [["jamones", "Jamones", "Piezas y loncheado"]]
    return _xlsx_response(_build_xlsx("Categorias", CATEGORY_COLS, sample), "plantilla_categorias.xlsx")


@router.post("/categories/export")
async def categories_export(payload: IdsIn, _=Depends(require_permission("products.read"))):
    items = await db.categories.find({}, {"_id": 0}).to_list(1000)
    rows = [[c.get("slug", ""), c.get("name", ""), c.get("description", "")] for c in items]
    return _xlsx_response(_build_xlsx("Categorias", CATEGORY_COLS, rows), "categorias_export.xlsx")


@router.post("/categories/import")
async def categories_import(file: UploadFile = File(...), _=Depends(require_permission("products.write"))):
    data = await file.read()
    _, rows = _read_xlsx(data)
    imported = updated = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    for idx, r in enumerate(rows, start=2):
        try:
            slug = str(r.get("Slug") or "").strip().lower()
            if not slug: continue
            doc = {"name": str(r.get("Nombre") or slug), "description": str(r.get("Descripción") or "")}
            existing = await db.categories.find_one({"slug": slug})
            if existing:
                await db.categories.update_one({"slug": slug}, {"$set": doc})
                updated += 1
            else:
                doc["id"] = str(uuid.uuid4())
                doc["slug"] = slug
                doc["created_at"] = now
                await db.categories.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {e}")
    return {"imported": imported, "updated": updated, "errors": errors}


# ============ USERS ============
USER_COLS = ["Email", "Nombre", "Rol", "Activo (SI/NO)", "Permisos (coma)", "Password (solo nuevo)"]


@router.get("/users/template")
async def users_template(_=Depends(require_permission("users.read"))):
    sample = [["nuevo@ejemplo.com", "Nombre Apellido", "manager", "SI", "products.read, orders.read", "Pass1234"]]
    return _xlsx_response(_build_xlsx("Usuarios", USER_COLS, sample), "plantilla_usuarios.xlsx")


@router.post("/users/export")
async def users_export(payload: IdsIn, _=Depends(require_permission("users.read"))):
    items = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(2000)
    rows = []
    for u in items:
        rows.append([
            u.get("email", ""), u.get("name", ""), u.get("role", ""),
            "SI" if u.get("is_active") else "NO",
            ", ".join(u.get("permissions") or ([] if not u.get("is_superadmin") else ALL_PERMISSIONS)),
            "",  # never export passwords
        ])
    return _xlsx_response(_build_xlsx("Usuarios", USER_COLS, rows), "usuarios_export.xlsx")


@router.post("/users/import")
async def users_import(file: UploadFile = File(...), _=Depends(require_permission("users.write"))):
    data = await file.read()
    _, rows = _read_xlsx(data)
    imported = updated = 0
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    for idx, r in enumerate(rows, start=2):
        try:
            email = str(r.get("Email") or "").strip().lower()
            if not email: continue
            perms_str = str(r.get("Permisos (coma)") or "")
            perms = [p.strip() for p in perms_str.split(",") if p.strip() in ALL_PERMISSIONS]
            doc = {
                "name": str(r.get("Nombre") or email),
                "role": str(r.get("Rol") or "manager"),
                "is_active": _bool(r.get("Activo (SI/NO)")) if r.get("Activo (SI/NO)") is not None else True,
                "permissions": perms,
                "updated_at": now,
            }
            existing = await db.users.find_one({"email": email})
            if existing:
                if existing.get("is_superadmin"):
                    errors.append(f"Fila {idx}: no se modifica el superadmin")
                    continue
                await db.users.update_one({"email": email}, {"$set": doc})
                updated += 1
            else:
                password = str(r.get("Password (solo nuevo)") or "")
                if not validate_password(password):
                    errors.append(f"Fila {idx}: password no válida (mín 8, mayús, minús, nº)")
                    continue
                doc["id"] = str(uuid.uuid4())
                doc["email"] = email
                doc["password_hash"] = hash_password(password)
                doc["is_superadmin"] = False
                doc["created_at"] = now
                await db.users.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {e}")
    return {"imported": imported, "updated": updated, "errors": errors}


# ============ CUSTOMERS ============
CUSTOMER_COLS = ["Email", "Nombre", "Teléfono", "NIF/CIF", "Direcciones", "Alta"]


def _addr_compact(a):
    parts = filter(None, [a.get("full_name"), a.get("address"), a.get("postal_code"), a.get("city"), a.get("country")])
    return ", ".join(parts)


@router.get("/customers/template")
async def customers_template(_=Depends(require_permission("users.read"))):
    sample = [["cliente@example.com", "Cliente Ejemplo", "+34 600 000 000", "12345678A", "", ""]]
    return _xlsx_response(_build_xlsx("Clientes", CUSTOMER_COLS, sample), "plantilla_clientes.xlsx")


@router.post("/customers/export")
async def customers_export(payload: IdsIn, _=Depends(require_permission("users.read"))):
    items = await db.customers.find({}, {"_id": 0, "password_hash": 0}).to_list(2000)
    rows = []
    for c in items:
        addrs = " | ".join(_addr_compact(a) for a in (c.get("addresses") or []))
        rows.append([
            c.get("email", ""), c.get("name", ""), c.get("phone", ""),
            c.get("tax_id", ""), addrs, (c.get("created_at") or "")[:10],
        ])
    return _xlsx_response(_build_xlsx("Clientes", CUSTOMER_COLS, rows), "clientes_export.xlsx")
