"""Excel import/export for orders.

- GET /api/excel/orders/template — returns a sample/blank Excel template
- POST /api/excel/orders/export — given an optional list of order_ids, returns an XLSX
- POST /api/excel/orders/import — accepts an XLSX upload, creates/updates orders
"""
import io
import uuid
from datetime import datetime, timezone
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from db import db
from auth import require_permission

router = APIRouter(prefix="/api/excel", tags=["excel"])


COLUMNS = [
    "Nº Pedido",
    "Nº Factura",
    "Fecha",
    "Cliente",
    "Email",
    "Teléfono",
    "NIF/CIF",
    "Dirección",
    "CP",
    "Ciudad",
    "País",
    "Productos (SKU x cantidad; …)",
    "Subtotal",
    "IVA",
    "Total",
    "Estado",
    "Estado pago",
    "Notas",
]


def _style_header(ws):
    header_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    header_font = Font(color="C5A059", bold=True, size=10)
    border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 26


def _autosize(ws):
    widths = [16, 16, 12, 24, 28, 14, 14, 30, 8, 18, 12, 50, 12, 12, 12, 16, 14, 30]
    for i, w in enumerate(widths, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = w


def _order_to_row(o: dict):
    cust = o.get("customer", {}) or {}
    items_str = "; ".join([f"{it.get('sku') or it.get('product_id')} x {it.get('qty')}" for it in (o.get("items") or [])])
    return [
        o.get("order_number", ""),
        o.get("invoice_number", ""),
        o.get("created_at", "")[:10],
        cust.get("name", ""),
        cust.get("email", ""),
        cust.get("phone", ""),
        cust.get("tax_id", ""),
        cust.get("address", ""),
        cust.get("postal_code", ""),
        cust.get("city", ""),
        cust.get("country", ""),
        items_str,
        float(o.get("subtotal", 0)),
        float(o.get("vat_total", 0)),
        float(o.get("total", 0)),
        o.get("status", ""),
        o.get("payment_status", ""),
        cust.get("notes", ""),
    ]


def _build_workbook(orders: list) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedidos"
    _style_header(ws)
    for i, o in enumerate(orders, start=2):
        row = _order_to_row(o)
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)
    _autosize(ws)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/orders/template")
async def template(_=Depends(require_permission("orders.read"))):
    # Build sample template (header + 1 sample row)
    sample = [{
        "order_number": "P-2026-00001",
        "invoice_number": "F-2026-00001",
        "created_at": "2026-02-12T10:00:00+00:00",
        "items": [
            {"sku": "JIB-LON-100", "qty": 2},
            {"sku": "EMB-CHO-001", "qty": 1},
        ],
        "subtotal": 27.09,
        "vat_total": 2.71,
        "total": 29.80,
        "status": "paid",
        "payment_status": "paid",
        "customer": {
            "name": "Cliente Ejemplo",
            "email": "cliente@example.com",
            "phone": "+34 600 000 000",
            "tax_id": "12345678A",
            "address": "Calle Ejemplo 1",
            "postal_code": "41230",
            "city": "Castilblanco de los Arroyos",
            "country": "España",
            "notes": "Entregar en horario de mañana",
        },
    }]
    pdf = _build_workbook(sample)
    return Response(
        content=pdf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_pedidos_las_dos_doncellas.xlsx"'},
    )


class ExportIn(BaseModel):
    order_ids: Optional[List[str]] = None
    status: Optional[str] = None


@router.post("/orders/export")
async def export_orders(payload: ExportIn, _=Depends(require_permission("orders.read"))):
    query = {}
    if payload.order_ids:
        query["id"] = {"$in": payload.order_ids}
    if payload.status:
        query["status"] = payload.status
    cursor = db.orders.find(query, {"_id": 0}).sort("created_at", -1)
    orders = [o async for o in cursor]
    if not orders:
        raise HTTPException(status_code=404, detail="No hay pedidos para exportar")
    xlsx = _build_workbook(orders)
    return Response(
        content=xlsx,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="pedidos_export.xlsx"'},
    )


@router.post("/orders/import")
async def import_orders(file: UploadFile = File(...), _=Depends(require_permission("orders.write"))):
    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo está vacío")

    imported = 0
    updated = 0
    errors: list[str] = []
    now = datetime.now(timezone.utc).isoformat()

    for idx, r in enumerate(rows, start=2):
        try:
            (order_number, invoice_number, fecha, cliente, email, tel, nif, direccion,
             cp, ciudad, pais, items_str, subtotal, iva, total, estado, estado_pago, notas) = (list(r) + [None] * 18)[:18]
            if not order_number:
                continue

            # Parse items "SKU x QTY; SKU x QTY"
            items = []
            for chunk in (items_str or "").split(";"):
                chunk = chunk.strip()
                if not chunk:
                    continue
                if "x" in chunk.lower():
                    parts = chunk.lower().split("x")
                    sku = parts[0].strip().upper()
                    qty = int(float(parts[1].strip())) if parts[1].strip() else 1
                else:
                    sku = chunk.upper()
                    qty = 1
                prod = await db.products.find_one({"sku": sku})
                items.append({
                    "product_id": prod["id"] if prod else None,
                    "sku": sku,
                    "name": prod["name"] if prod else sku,
                    "qty": qty,
                    "unit_price": float(prod["price"]) if prod else 0.0,
                    "base_unit_price": 0.0,
                    "vat_rate": int(prod.get("vat_rate", 10)) if prod else 10,
                    "line_total": float(prod["price"]) * qty if prod else 0.0,
                    "image": (prod.get("images") or [None])[0] if prod else None,
                })

            customer = {
                "name": cliente or "",
                "email": email or "",
                "phone": tel or "",
                "tax_id": nif or "",
                "address": direccion or "",
                "postal_code": str(cp) if cp is not None else "",
                "city": ciudad or "",
                "country": pais or "España",
                "notes": notas or "",
            }

            existing = await db.orders.find_one({"order_number": order_number})
            doc_set = {
                "invoice_number": invoice_number or f"F-IMP-{order_number}",
                "invoice_date": (str(fecha)[:10] if fecha else now[:10]),
                "items": items,
                "customer": customer,
                "subtotal": float(subtotal or 0),
                "vat_total": float(iva or 0),
                "vat_breakdown": {},
                "shipping": 0.0,
                "total": float(total or 0),
                "currency": "eur",
                "status": estado or "paid",
                "payment_status": estado_pago or "paid",
                "updated_at": now,
            }
            if existing:
                await db.orders.update_one({"order_number": order_number}, {"$set": doc_set})
                updated += 1
            else:
                doc = {
                    "id": str(uuid.uuid4()),
                    "order_number": order_number,
                    "session_id": None,
                    "tracking": [{"status": doc_set["status"], "at": now, "note": "Importado desde Excel"}],
                    "created_at": now,
                    **doc_set,
                }
                await db.orders.insert_one(doc)
                imported += 1
        except Exception as e:
            errors.append(f"Fila {idx}: {e}")

    return {"imported": imported, "updated": updated, "errors": errors}
